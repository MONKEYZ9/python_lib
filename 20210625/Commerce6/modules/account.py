import openpyxl
from .error_catch import ErrorDetect as err #에러 모듈

class Commerce_login :
  
  def __init__(self) :
    self.id = str()
    self.pw = str()
    self.nickname = str()

  def login(self, login_id, login_pw) :
    load_wb = openpyxl.load_workbook(r'./database/login/login.xlsx', data_only = True) #엑셀 불러오기
    load_ws = load_wb['Sheet'] #엑셀의 시트명으로 특정 시트 불러오기
    max_row_loca = load_ws.max_row

    member_info_collection = load_ws['A2' : f'C{max_row_loca}']
    member_info_dict = {}

    for id, pw, nick in member_info_collection: #각 cell의 값들을 저장하기 위함
        member_info_dict[id.value] = [pw.value, nick.value]
    
    if login_id in member_info_dict:
        
        if login_pw == member_info_dict[login_id][0]: #아이디, pw 둘 다 성공
            
            user_nick = member_info_dict[login_id][1]
            self.id = login_id
            self.pw = login_pw
            self.nickname = user_nick
            return 'complete'

        else:
            return 'PwFailedUser'

    return 'IdFailedUser'

  def new_account(self) :
    load_wb = openpyxl.load_workbook(r'./database/login/login.xlsx', data_only = True) #엑셀 불러오기
    load_ws = load_wb['Sheet'] #엑셀의 시트명으로 특정 시트 불러오기

    while True:
      condition = True #중복값 제거용도
      new_id = input('사용하실 아이디를 입력하세요 : ')

      if new_id.strip() == '': #스페이스 등 빈 값 들어올 경우 막음
        print(err.error_alert('EmptyValue'))
        continue
      
      else:
        for idname in load_ws['A']: #A열(id가 있는 열)의 값들을 다 검색
          if new_id == idname.value: 
            print(err.error_alert('IdOverlapped')) #기존 아이디 있으면, 오류 alert
            condition = False
            break

      if condition == True:
        break


    while True:
      new_pw = input('사용하실 패스워드를 입력하세요 : ')

      if new_pw.strip() == '':
        print(err.error_alert('EmptyValue'))
        continue
      
      pw_tmp = input('비밀번호 확인. 다시 한 번 패스워드를 입력하세요 : ')

      if new_pw != pw_tmp:
          print('비밀번호가 일치하지 않습니다. 다시 입력하세요.', end='\n\n')
          continue
      
      break
            
    nickname = input('사용하실 닉네임을 입력해주세요 : ')
    self.id = new_id
    self.pw = new_pw
    self.nickname = nickname
    user_info = [new_id, new_pw, nickname]
    max_row_loca = load_ws.max_row + 1 #기존 자료에 덮어쓰지 않게 마지막줄 + 1

    for col in range(1,4):
        load_ws.cell(row = max_row_loca, column = col, value = user_info[col - 1]) #id, pwd, nickname을 채우고있음

    load_wb.save(r'./database/login/login.xlsx')
    load_wb.close()

    print('회원가입이 완료되었습니다!', end='\n\n')
    return 'complete'
