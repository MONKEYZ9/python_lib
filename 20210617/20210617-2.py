# -*- coding: utf-8 -*-

# LV2 : excel / 엑셀 파일을 불러와주세요.
# step2.xlsx

# 결과
# +hoodies 1072
# +hoody 990
# Google Merchandise Store 459
# +sweatshirt 566
# youtuber merch 345
# +tumbler 276
# (not set) 123
# hoodies 247
# +youtube +merchandise 203
# Google Apparel 110
# +youtubers +merch 128
# +youtuber +merch 120
# YouTube Merchandise Store 107
import openpyxl
from pathlib import Path

# xlsx_file = Path('C://Users\mingzzang\Desktop\step2.xlsx')
with open('C://Users\mingzzang\Desktop\step2.xlsx') as xlsx_file:
    obj = openpyxl.load_workbook(xlsx_file)
    sheet = obj.active
    for row in sheet.iter_rows(max_row=14):
        for cell in row:
            print(cell.value, end=" ")
        print()
# with open('C://Users\mingzzang\Desktop\step2.xlsx') as excel_file:
#     excel_read = openpyxl.load_workbook(excel_file.name)
    #  excel_read = openpyxl.load_workbook(excel_fil) 여기까지 하면 이거는 절대경로를 가져오는거야
    # <openpyxl.workbook.workbook.Workbook object at 0x0000020A7A2353A0>
    # 그래서 name을 가져와서 한다는 거야
    # print(excel_read)
    # excel_sheet_read = excel_read.active #시트를 인식시켰고
    # for row in excel_sheet_read.iter_rows(max_row=14): # row를 가져올 수 있게 하는데 
    # for row in excel_sheet_read.rows:
    #     print(row[0].value, row[1].value)
        # for cell in row:
        #     print(cell.value, end=" ")
        # print()